from __future__ import annotations

from pathlib import Path
from typing import Dict, List

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from .database import SessionLocal
from . import models


XLSX_PATH = Path("Kits_Ferramental_V1.xlsx")


def read_sheet_rows(xlsx_path: Path, sheet_name: str) -> List[Dict[str, object]]:
    wb = load_workbook(xlsx_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet nao encontrada: {sheet_name}")

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    data = []

    for row in rows[1:]:
        if not row:
            continue
        record = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            if idx < len(row):
                record[header] = row[idx]
        if record:
            data.append(record)

    return data


def get_or_create_setor(db: Session, nome: str) -> models.Setor:
    nome = nome.strip()
    setor = db.query(models.Setor).filter(models.Setor.nome == nome).first()
    if not setor:
        setor = models.Setor(nome=nome)
        db.add(setor)
        db.commit()
        db.refresh(setor)
    return setor


def get_or_create_kit(db: Session, nome: str, setor_id: int) -> models.Kit:
    kit = (
        db.query(models.Kit)
        .filter(models.Kit.nome == nome)
        .filter(models.Kit.setor_id == setor_id)
        .first()
    )
    if not kit:
        kit = models.Kit(nome=nome, setor_id=setor_id)
        db.add(kit)
        db.commit()
        db.refresh(kit)
    return kit


def get_item_by_patrimonio(db: Session, patrimonio: str):
    return (
        db.query(models.Item)
        .filter(models.Item.patrimonio == patrimonio.strip())
        .first()
    )


def get_or_create_kit_item(db: Session, kit_id: int, item_id: int) -> bool:
    existing = (
        db.query(models.KitItem)
        .filter(models.KitItem.kit_id == kit_id)
        .filter(models.KitItem.item_id == item_id)
        .first()
    )
    if existing:
        return False
    rel = models.KitItem(kit_id=kit_id, item_id=item_id, quantidade=1)
    db.add(rel)
    return True


def importar_kits(xlsx_path: Path = XLSX_PATH) -> None:
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Arquivo nao encontrado: {xlsx_path.resolve()}")

    print("\n>>> IMPORTANDO KITS/ITENS (XLSX)")
    db = SessionLocal()

    try:
        classificacao = read_sheet_rows(xlsx_path, "01_Itens_Classificados")
        kits_rows = read_sheet_rows(xlsx_path, "02_Kits")
        kit_itens = read_sheet_rows(xlsx_path, "03_Kit_Itens")

        setores_unicos = sorted(
            {
                str(r.get("SETOR", "")).strip()
                for r in classificacao
                if str(r.get("SETOR", "")).strip()
            }
        )

        mapa_setores: Dict[str, int] = {}
        for nome_setor in setores_unicos:
            setor = get_or_create_setor(db, nome_setor)
            mapa_setores[nome_setor] = setor.id

        print(f"Setores importados: {len(mapa_setores)}")

        mapa_kits: Dict[object, int] = {}
        for row in kits_rows:
            nome = str(row.get("NomeKit", "")).strip()
            setor_nome = str(row.get("Setor", "")).strip()
            if not nome or not setor_nome:
                continue
            setor_id = mapa_setores.get(setor_nome)
            if not setor_id:
                continue
            kit = get_or_create_kit(db, nome, setor_id)
            mapa_kits[row.get("KitID")] = kit.id

        print(f"Kits importados: {len(mapa_kits)}")

        count_rel = 0
        for row in kit_itens:
            kit_id_ref = row.get("KitID")
            patrimonio = str(row.get("IDENTIFICACAO", "")).strip()
            if not kit_id_ref or not patrimonio:
                continue
            kit_id_real = mapa_kits.get(kit_id_ref)
            if not kit_id_real:
                continue
            item = get_item_by_patrimonio(db, patrimonio)
            if not item:
                continue
            if get_or_create_kit_item(db, kit_id_real, item.id):
                count_rel += 1

        db.commit()
        print(f"Itens alocados nos kits: {count_rel}")
        print("\nIMPORTACAO FINALIZADA COM SUCESSO!\n")
    finally:
        db.close()


if __name__ == "__main__":
    importar_kits()
