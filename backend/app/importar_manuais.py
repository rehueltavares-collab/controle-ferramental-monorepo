from __future__ import annotations

import argparse
from datetime import date
from pathlib import Path
from typing import Dict, Iterable, Tuple

from openpyxl import load_workbook
from sqlalchemy import text
from sqlalchemy.orm import Session

from .database import SessionLocal


def norm_text(s: object) -> str:
    return str(s or "").strip()


def find_header_row(rows: Iterable[Tuple[object, ...]]) -> Tuple[int, Dict[str, int]]:
    for idx, row in enumerate(rows, start=1):
        vals = [norm_text(c).upper() for c in row]
        if not vals:
            continue
        if "COLUNA 1" in vals and "MATERIAL" in vals:
            return idx, {v: i for i, v in enumerate(vals)}
    raise RuntimeError("Cabecalho nao encontrado (COLUNA 1/MATERIAL).")


def get_or_create_manual_item(db: Session, nome: str) -> int:
    row = db.execute(
        text("SELECT id FROM manual_itens WHERE nome = :n LIMIT 1"),
        {"n": nome},
    ).first()
    if row:
        return int(row[0])
    db.execute(text("INSERT INTO manual_itens (nome, ativo) VALUES (:n, 1)"), {"n": nome})
    db.commit()
    row = db.execute(
        text("SELECT id FROM manual_itens WHERE nome = :n LIMIT 1"),
        {"n": nome},
    ).first()
    return int(row[0])


def find_subresponsavel_id(db: Session, nome: str) -> int | None:
    row = db.execute(
        text(
            """
            SELECT id FROM subresponsaveis
            WHERE nome = :n
            LIMIT 1
            """
        ),
        {"n": nome},
    ).first()
    return int(row[0]) if row else None


def upsert_posse(
    db: Session,
    sub_id: int,
    manual_item_id: int,
    quantidade: int,
    data_retirada: date | None,
) -> None:
    row = db.execute(
        text(
            """
            SELECT id FROM manual_posse
            WHERE subresponsavel_id = :sub_id AND manual_item_id = :item_id
            LIMIT 1
            """
        ),
        {"sub_id": sub_id, "item_id": manual_item_id},
    ).first()
    if row:
        db.execute(
            text(
                """
                UPDATE manual_posse
                SET quantidade = :q, data_retirada = :d
                WHERE id = :id
                """
            ),
            {"q": quantidade, "d": data_retirada, "id": int(row[0])},
        )
    else:
        db.execute(
            text(
                """
                INSERT INTO manual_posse
                (subresponsavel_id, manual_item_id, quantidade, data_retirada)
                VALUES
                (:sub_id, :item_id, :q, :d)
                """
            ),
            {"sub_id": sub_id, "item_id": manual_item_id, "q": quantidade, "d": data_retirada},
        )


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True, help="Caminho do .xlsx")
    ap.add_argument("--sheet", default=None, help="Nome da aba (opcional).")
    ap.add_argument("--dry", action="store_true", help="Nao grava no banco.")
    args = ap.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        raise SystemExit(f"[ERRO] Arquivo nao encontrado: {xlsx_path}")

    wb = load_workbook(xlsx_path, data_only=True)
    sheet_name = args.sheet or wb.sheetnames[0]
    if sheet_name not in wb.sheetnames:
        raise SystemExit(f"[ERRO] Aba nao encontrada: {sheet_name}")

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    header_row_idx, header_map = find_header_row(rows)

    nome_idx = header_map.get("COLUNA 1")
    status_idx = header_map.get("STATUS")
    material_idx = header_map.get("MATERIAL")
    qtd_idx = header_map.get("QNTD")
    data_idx = header_map.get("DATA DE RETIRADA")

    if nome_idx is None or material_idx is None:
        raise SystemExit("[ERRO] Colunas obrigatorias nao encontradas.")

    db = SessionLocal()
    try:
        total = 0
        itens_criados = 0
        posse_criada = 0
        posse_atualizada = 0
        sem_sub = 0

        for row in rows[header_row_idx:]:
            total += 1
            nome = norm_text(row[nome_idx]) if nome_idx < len(row) else ""
            status = norm_text(row[status_idx]) if status_idx is not None and status_idx < len(row) else ""
            material = norm_text(row[material_idx]) if material_idx < len(row) else ""
            qtd = row[qtd_idx] if qtd_idx is not None and qtd_idx < len(row) else 1
            data_raw = row[data_idx] if data_idx is not None and data_idx < len(row) else None

            if not nome or not material:
                continue
            if status and status.upper() != "ATIVO":
                continue

            manual_item_id = get_or_create_manual_item(db, material)

            sub_id = find_subresponsavel_id(db, nome)
            if not sub_id:
                sem_sub += 1
                continue

            quantidade = int(qtd) if isinstance(qtd, (int, float)) else 1
            data_retirada = data_raw.date() if hasattr(data_raw, "date") else None

            row_before = db.execute(
                text(
                    """
                    SELECT id FROM manual_posse
                    WHERE subresponsavel_id = :sub_id AND manual_item_id = :item_id
                    LIMIT 1
                    """
                ),
                {"sub_id": sub_id, "item_id": manual_item_id},
            ).first()

            if not args.dry:
                upsert_posse(db, sub_id, manual_item_id, quantidade, data_retirada)
                db.commit()

            if row_before:
                posse_atualizada += 1
            else:
                posse_criada += 1
        if not args.dry:
            db.commit()

        print("IMPORTACAO MANUAIS FINALIZADA")
        print(f"Total linhas lidas: {total}")
        print(f"Posse criada: {posse_criada}")
        print(f"Posse atualizada: {posse_atualizada}")
        print(f"Sem subresponsavel correspondente: {sem_sub}")
    finally:
        db.close()


if __name__ == "__main__":
    main()
