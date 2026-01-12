from __future__ import annotations

import argparse
import re
import unicodedata
from pathlib import Path
from typing import Dict, Iterable, Tuple

from openpyxl import load_workbook

from .database import SessionLocal
from . import models


def only_digits(s: str) -> str:
    return re.sub(r"\D+", "", (s or "").strip())


def norm_text(s: object) -> str:
    return str(s or "").strip()


def norm_header(s: object) -> str:
    raw = norm_text(s).upper()
    return "".join(
        ch for ch in unicodedata.normalize("NFKD", raw) if not unicodedata.combining(ch)
    )


def find_header_row(rows: Iterable[Tuple[object, ...]]) -> Tuple[int, Dict[str, int]]:
    for idx, row in enumerate(rows, start=1):
        row_vals = [norm_header(c) for c in row]
        if not row_vals:
            continue
        if "AREA" in row_vals and "FUNCAO" in row_vals and "RESPONSAVEL" in row_vals:
            return idx, {v: i for i, v in enumerate(row_vals)}
        if "NOME" in row_vals and "FUNCAO" in row_vals and "OBRA" in row_vals:
            return idx, {v: i for i, v in enumerate(row_vals)}
    raise RuntimeError("Cabecalho nao encontrado (AREA/FUNCAO/RESPONSAVEL).")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True, help="Caminho do .xlsx")
    ap.add_argument("--sheet", default=None, help="Nome da aba (opcional). Se vazio, usa a 1a.")
    ap.add_argument("--dry", action="store_true", help="Nao grava no banco (so simula).")
    args = ap.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        raise SystemExit(f"[ERRO] Arquivo nao encontrado: {xlsx_path}")

    wb = load_workbook(xlsx_path, data_only=True)
    sheet_name = args.sheet or wb.sheetnames[0]
    if sheet_name not in wb.sheetnames:
        raise SystemExit(f"[ERRO] Aba nao encontrada: {sheet_name}")

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    header_row_idx, header_map = find_header_row(rows)

    area_idx = header_map.get("AREA")
    funcao_idx = header_map.get("FUNCAO")
    resp_idx = header_map.get("RESPONSAVEL")
    contato_idx = header_map.get("CONTATO")

    if area_idx is None and "OBRA" in header_map:
        area_idx = header_map.get("OBRA")
    if resp_idx is None and "NOME" in header_map:
        resp_idx = header_map.get("NOME")

    if area_idx is None or funcao_idx is None or resp_idx is None:
        raise SystemExit("[ERRO] Colunas obrigatorias nao encontradas.")

    db = SessionLocal()
    try:
        setores_criados = 0
        encarregados_criados = 0
        encarregados_pulados = 0

        setor_cache: Dict[str, int] = {}

        for row in rows[header_row_idx:]:
            area = norm_text(row[area_idx]) if area_idx < len(row) else ""
            funcao = norm_text(row[funcao_idx]) if funcao_idx < len(row) else ""
            nome = norm_text(row[resp_idx]) if resp_idx < len(row) else ""
            tel = norm_text(row[contato_idx]) if contato_idx is not None and contato_idx < len(row) else ""

            if not area or not nome:
                continue

            if area not in setor_cache:
                setor = db.query(models.Setor).filter(models.Setor.nome == area).first()
                if not setor:
                    setor = models.Setor(nome=area)
                    db.add(setor)
                    if not args.dry:
                        db.commit()
                        db.refresh(setor)
                    setores_criados += 1
                setor_cache[area] = setor.id

            setor_id = setor_cache[area]

            exists = (
                db.query(models.Encarregado)
                .filter(models.Encarregado.setor_id == setor_id)
                .filter(models.Encarregado.nome == nome)
                .first()
            )
            if exists:
                encarregados_pulados += 1
                continue

            novo = models.Encarregado(
                setor_id=setor_id,
                funcao=funcao or "Encarregado",
                nome=nome,
                telefone=only_digits(tel),
            )
            db.add(novo)
            if not args.dry:
                db.commit()
                db.refresh(novo)
            encarregados_criados += 1

        print("IMPORTACAO DE ENCARREGADOS FINALIZADA")
        print(f"Setores criados (ou novos encontrados): {setores_criados}")
        print(f"Encarregados criados: {encarregados_criados}")
        print(f"Encarregados pulados (ja existiam): {encarregados_pulados}")
    finally:
        db.close()


if __name__ == "__main__":
    main()
