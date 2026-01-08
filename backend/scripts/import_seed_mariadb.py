# backend/scripts/import_seed_mariadb.py
from __future__ import annotations

import os
import re
import csv
import sqlite3
from pathlib import Path
from datetime import datetime

import pymysql

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None


# =========================
# Paths / .env loader (simples, sem depender de python-dotenv)
# =========================
def project_root() -> Path:
    # .../backend/scripts/import_seed_mariadb.py -> parents[2] = .../controle-ferramental-monorepo
    return Path(__file__).resolve().parents[2]

def load_env_file(env_path: Path) -> None:
    if not env_path.exists():
        return
    for line in env_path.read_text(encoding="utf-8", errors="ignore").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        k = k.strip()
        v = v.strip().strip('"').strip("'")
        os.environ.setdefault(k, v)

ROOT = project_root()
load_env_file(ROOT / "backend" / ".env")


# =========================
# MariaDB connection
# =========================
def mariadb_conn():
    host = os.getenv("DB_HOST", "localhost")
    port = int(os.getenv("DB_PORT", "3306"))
    user = os.getenv("DB_USER", "ferramental_app")
    password = os.getenv("DB_PASSWORD", "123456")
    db = os.getenv("DB_NAME", "ferramental")

    return pymysql.connect(
        host=host,
        port=port,
        user=user,
        password=password,
        database=db,
        charset="utf8mb4",
        autocommit=False,
        cursorclass=pymysql.cursors.DictCursor,
    )

def ensure_schema(con):
    with con.cursor() as cur:
        # subresponsáveis
        cur.execute("""
        CREATE TABLE IF NOT EXISTS subresponsaveis (
          id INT PRIMARY KEY,
          nome VARCHAR(255) NOT NULL,
          secao VARCHAR(120) NULL,
          ativo TINYINT NOT NULL DEFAULT 1,
          pin CHAR(6) NULL,
          INDEX idx_nome (nome),
          INDEX idx_secao (secao)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
        """)

        # itens manuais
        cur.execute("""
        CREATE TABLE IF NOT EXISTS manual_itens (
          id INT AUTO_INCREMENT PRIMARY KEY,
          nome VARCHAR(255) NOT NULL,
          ativo TINYINT NOT NULL DEFAULT 1,
          UNIQUE KEY uq_manual_item_nome (nome)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
        """)

        # posse manuais (sem id por item no mundo real → aqui vira item normalizado + quantidade)
        cur.execute("""
        CREATE TABLE IF NOT EXISTS manual_posse (
          id INT AUTO_INCREMENT PRIMARY KEY,
          subresponsavel_id INT NOT NULL,
          manual_item_id INT NOT NULL,
          quantidade INT NOT NULL DEFAULT 1,
          data_retirada DATE NULL,
          atualizado_em TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
          UNIQUE KEY uq_posse (subresponsavel_id, manual_item_id),
          CONSTRAINT fk_posse_sub FOREIGN KEY (subresponsavel_id) REFERENCES subresponsaveis(id),
          CONSTRAINT fk_posse_item FOREIGN KEY (manual_item_id) REFERENCES manual_itens(id)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
        """)

        # itens elétricos (patrimônio = identificação)
        cur.execute("""
        CREATE TABLE IF NOT EXISTS itens_eletricos (
          id INT AUTO_INCREMENT PRIMARY KEY,
          patrimonio VARCHAR(80) NOT NULL,
          descricao VARCHAR(255) NOT NULL,
          ativo TINYINT NOT NULL DEFAULT 1,
          UNIQUE KEY uq_patrimonio (patrimonio),
          INDEX idx_desc (descricao)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
        """)

        # movimentos (auditoria geral)
        cur.execute("""
        CREATE TABLE IF NOT EXISTS movimentos (
          id INT AUTO_INCREMENT PRIMARY KEY,
          tipo ENUM('DISTRIBUIR','RECOLHER','MANUAL_ENTREGAR','MANUAL_DEVOLVER') NOT NULL,
          kit_id INT NULL,
          patrimonio VARCHAR(80) NULL,
          encarregado_id INT NULL,
          subresponsavel_id INT NULL,
          manual_item_id INT NULL,
          quantidade INT NULL,
          created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
          observacao VARCHAR(255) NULL,
          INDEX idx_mov_sub (subresponsavel_id),
          INDEX idx_mov_tipo (tipo),
          INDEX idx_mov_patrimonio (patrimonio)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
        """)

    con.commit()

def upsert_subresponsavel(con, row: dict):
    with con.cursor() as cur:
        cur.execute("""
        INSERT INTO subresponsaveis (id, nome, secao, ativo, pin)
        VALUES (%s, %s, %s, %s, %s)
        ON DUPLICATE KEY UPDATE
          nome=VALUES(nome),
          secao=VALUES(secao),
          ativo=VALUES(ativo),
          pin=VALUES(pin);
        """, (row["id"], row["nome"], row["secao"], row["ativo"], row["pin"]))

def get_or_create_manual_item(con, nome: str) -> int:
    nome = (nome or "").strip()
    if not nome:
        raise ValueError("manual item vazio")

    with con.cursor() as cur:
        cur.execute("SELECT id FROM manual_itens WHERE nome=%s", (nome,))
        r = cur.fetchone()
        if r:
            return int(r["id"])

        cur.execute("INSERT INTO manual_itens (nome, ativo) VALUES (%s, 1)", (nome,))
        return int(cur.lastrowid)

def upsert_manual_posse(con, sub_id: int, item_id: int, qtd: int, data_retirada):
    with con.cursor() as cur:
        cur.execute("""
        INSERT INTO manual_posse (subresponsavel_id, manual_item_id, quantidade, data_retirada)
        VALUES (%s, %s, %s, %s)
        ON DUPLICATE KEY UPDATE
          quantidade=VALUES(quantidade),
          data_retirada=COALESCE(VALUES(data_retirada), data_retirada);
        """, (sub_id, item_id, int(qtd), data_retirada))

def upsert_item_eletrico(con, patrimonio: str, descricao: str):
    patrimonio = (patrimonio or "").strip()
    descricao = (descricao or "").strip()
    if not patrimonio or not descricao:
        return

    with con.cursor() as cur:
        cur.execute("""
        INSERT INTO itens_eletricos (patrimonio, descricao, ativo)
        VALUES (%s, %s, 1)
        ON DUPLICATE KEY UPDATE
          descricao=VALUES(descricao),
          ativo=1;
        """, (patrimonio, descricao))


# =========================
# SQLite → MariaDB (subresponsáveis)
# =========================
def sqlite_path() -> Path:
    p = os.getenv("SQLITE_PATH", "backend/data/ferramental.db")
    return (ROOT / p).resolve()

def import_subresponsaveis_from_sqlite(con):
    src = sqlite_path()
    if not src.exists():
        print(f"[WARN] SQLite não encontrado em: {src}")
        return 0

    s = sqlite3.connect(str(src))
    s.row_factory = sqlite3.Row
    cur = s.cursor()

    rows = cur.execute("""
        SELECT id, nome, secao, ativo, pin
        FROM subresponsaveis
        ORDER BY id
    """).fetchall()

    total = 0
    for r in rows:
        row = {
            "id": int(r["id"]),
            "nome": (r["nome"] or "").strip(),
            "secao": (r["secao"] or "").strip() or None,
            "ativo": int(r["ativo"] or 0),
            "pin": (r["pin"] or "").strip() or None,
        }
        if not row["nome"]:
            continue
        upsert_subresponsavel(con, row)
        total += 1

    con.commit()
    s.close()
    print(f"[OK] subresponsaveis migrados: {total}")
    return total


# =========================
# CSV (itens elétricos)
# =========================
def csv_eletricos_path() -> Path:
    p = os.getenv("CSV_ELETRICOS", "data/fontes/itens_ferramental_utf8.csv")
    return (ROOT / p).resolve()

def import_eletricos_from_csv(con):
    src = csv_eletricos_path()
    if not src.exists():
        print(f"[WARN] CSV eletricos não encontrado em: {src}")
        return 0

    # separador é ;
    with src.open("r", encoding="utf-8", errors="ignore", newline="") as f:
        reader = csv.DictReader(f, delimiter=";")
        # esperado: Identificação;Equipamento
        total = 0
        for row in reader:
            patrimonio = (row.get("Identificação") or row.get("Identificacao") or row.get("ID") or "").strip()
            descricao = (row.get("Equipamento") or row.get("Descricao") or row.get("Descrição") or "").strip()
            if not patrimonio or not descricao:
                continue
            upsert_item_eletrico(con, patrimonio, descricao)
            total += 1

    con.commit()
    print(f"[OK] itens_eletricos importados: {total}")
    return total


# =========================
# XLSX (manuais + posse)
# Parser flexível: tenta detectar colunas e itens
# =========================
def xlsx_manuais_path() -> Path:
    p = os.getenv("XLSX_MANUAIS", "data/fontes/termo_responsabilidade_almox_2024.xlsx")
    return (ROOT / p).resolve()

def _norm(s: str) -> str:
    s = (s or "").strip().upper()
    s = re.sub(r"\s+", " ", s)
    return s

def _parse_date(val):
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val.date()
    # tenta dd/mm/yyyy
    try:
        t = str(val).strip()
        if re.match(r"^\d{1,2}/\d{1,2}/\d{2,4}$", t):
            d, m, y = t.split("/")
            if len(y) == 2:
                y = "20" + y
            return datetime(int(y), int(m), int(d)).date()
    except Exception:
        return None
    return None

def import_manuais_from_xlsx(con):
    src = xlsx_manuais_path()
    if not src.exists():
        print(f"[WARN] XLSX manuais não encontrado em: {src}")
        return 0

    if load_workbook is None:
        raise RuntimeError("openpyxl não instalado. Rode: pip install openpyxl")

    wb = load_workbook(filename=str(src), data_only=True)

    total_posse = 0
    sheets = wb.sheetnames
    print(f"[INFO] planilhas detectadas: {sheets}")

    for sh_name in sheets:
        ws = wb[sh_name]
        # tenta achar header nas primeiras 20 linhas
        header_row_idx = None
        header = []

        for r in range(1, 21):
            vals = [ws.cell(r, c).value for c in range(1, 60)]
            joined = " | ".join([_norm(str(v)) for v in vals if v is not None])
            # heurística: linha que contenha NOME e algo de DATA
            if "NOME" in joined and ("DATA" in joined or "RETIRADA" in joined or "ENTREGA" in joined):
                header_row_idx = r
                header = [ws.cell(r, c).value for c in range(1, 60)]
                break

        if not header_row_idx:
            # se não achou, pula
            continue

        # mapeia colunas
        col_map = {}
        for idx, v in enumerate(header, start=1):
            key = _norm(str(v)) if v is not None else ""
            if not key:
                continue
            if "NOME" in key and "ITEM" not in key:
                col_map["nome"] = idx
            elif "SECAO" in key or "SEÇÃO" in key or "SETOR" in key:
                col_map["secao"] = idx
            elif "DATA" in key:
                col_map["data"] = idx

        if "nome" not in col_map:
            continue

        # itens: colunas que não são nome/secao/data e têm header preenchido
        item_cols = []
        for idx, v in enumerate(header, start=1):
            if idx in col_map.values():
                continue
            name = (str(v).strip() if v is not None else "")
            if not name:
                continue
            # ignora colunas típicas de assinatura/observação
            up = _norm(name)
            if any(x in up for x in ["ASSIN", "CPF", "RG", "MATR", "OBS", "TELEFONE"]):
                continue
            item_cols.append((idx, name.strip()))

        if not item_cols:
            continue

        print(f"[INFO] {sh_name}: header linha {header_row_idx} | itens detectados: {len(item_cols)}")

        # varre linhas seguintes
        for r in range(header_row_idx + 1, ws.max_row + 1):
            nome = ws.cell(r, col_map["nome"]).value
            if nome is None:
                continue
            nome = str(nome).strip()
            if not nome or len(nome) < 3:
                continue

            secao = str(ws.cell(r, col_map.get("secao", 0)).value).strip() if col_map.get("secao") else None
            data_retirada = _parse_date(ws.cell(r, col_map.get("data", 0)).value) if col_map.get("data") else None

            # tenta localizar subresponsável no DB (por nome exato). Se não achar, pula (não inventa)
            with con.cursor() as cur:
                cur.execute("SELECT id FROM subresponsaveis WHERE nome=%s LIMIT 1", (nome,))
                srow = cur.fetchone()

            if not srow:
                continue

            sub_id = int(srow["id"])

            # Para cada item, se a célula tiver algo que pareça quantidade/“X”
            for cidx, item_name in item_cols:
                val = ws.cell(r, cidx).value
                if val is None or val == "":
                    continue

                # normaliza quantidade
                qtd = 1
                try:
                    # se vier número
                    if isinstance(val, (int, float)):
                        qtd = int(val)
                    else:
                        t = str(val).strip()
                        # se for "X" ou "OK" considera 1
                        if re.match(r"^(X|OK|SIM|S)$", _norm(t)):
                            qtd = 1
                        else:
                            # tenta extrair número
                            m = re.search(r"\d+", t)
                            if m:
                                qtd = int(m.group(0))
                            else:
                                qtd = 1
                except Exception:
                    qtd = 1

                item_id = get_or_create_manual_item(con, item_name)
                upsert_manual_posse(con, sub_id, item_id, qtd, data_retirada)
                total_posse += 1

    con.commit()
    print(f"[OK] registros manual_posse upsertados: {total_posse}")
    return total_posse


def main():
    print("== Seed MariaDB Ferramental ==")
    print(f"[INFO] project_root: {ROOT}")

    con = mariadb_conn()
    try:
        ensure_schema(con)

        # 1) subresponsáveis (SQLite -> MariaDB)
        import_subresponsaveis_from_sqlite(con)

        # 2) itens elétricos (CSV)
        import_eletricos_from_csv(con)

        # 3) itens manuais + posse (XLSX)
        import_manuais_from_xlsx(con)

        con.commit()
        print("[DONE] seed finalizado com sucesso.")
    except Exception as e:
        con.rollback()
        print("[ERRO] seed falhou:", repr(e))
        raise
    finally:
        con.close()

if __name__ == "__main__":
    main()
