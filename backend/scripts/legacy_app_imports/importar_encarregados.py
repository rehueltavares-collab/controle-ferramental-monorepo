# importar_encarregados.py
# Importa encarregados a partir do Excel contatos_lideres_marica.xlsx
# e grava no banco ferramental.db (na raiz do repositório).
#
# Rodar a partir da pasta backend (recomendado):
#   cd backend
#   python ..\importar_encarregados.py

from __future__ import annotations

import re
import sqlite3
from pathlib import Path

from openpyxl import load_workbook


def norm(s: str) -> str:
    s = (s or "").strip().lower()
    # normalização simples (sem dependências)
    s = (
        s.replace("ç", "c")
        .replace("á", "a").replace("à", "a").replace("ã", "a").replace("â", "a")
        .replace("é", "e").replace("ê", "e")
        .replace("í", "i")
        .replace("ó", "o").replace("ô", "o").replace("õ", "o")
        .replace("ú", "u")
    )
    s = re.sub(r"\s+", " ", s)
    return s


def pick_col(headers: list[str], candidates: list[str]) -> int | None:
    nheaders = [norm(h) for h in headers]
    for cand in candidates:
        candn = norm(cand)
        for i, h in enumerate(nheaders):
            if candn == h or candn in h:
                return i
    return None


def ensure_schema(con: sqlite3.Connection) -> None:
    cur = con.cursor()
    # setores
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS setores (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT NOT NULL UNIQUE
        );
        """
    )
    # encarregados
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS encarregados (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            setor_id INTEGER NOT NULL,
            funcao TEXT,
            nome TEXT NOT NULL,
            telefone TEXT,
            FOREIGN KEY(setor_id) REFERENCES setores(id)
        );
        """
    )
    con.commit()


def get_or_create_setor(con: sqlite3.Connection, nome: str) -> int:
    nome = (nome or "").strip()
    if not nome:
        nome = "SEM SETOR"
    cur = con.cursor()
    row = cur.execute("SELECT id FROM setores WHERE nome = ?", (nome,)).fetchone()
    if row:
        return int(row[0])
    cur.execute("INSERT INTO setores (nome) VALUES (?)", (nome,))
    con.commit()
    return int(cur.lastrowid)


def exists_encarregado(con: sqlite3.Connection, setor_id: int, nome: str) -> bool:
    cur = con.cursor()
    row = cur.execute(
        "SELECT 1 FROM encarregados WHERE setor_id=? AND nome=? LIMIT 1",
        (setor_id, nome),
    ).fetchone()
    return bool(row)


def main() -> None:
    repo_root = Path(__file__).resolve().parent
    db_path = repo_root / "ferramental.db"
    xlsx_path = repo_root / "contatos_lideres_marica.xlsx"

    if not db_path.exists():
        raise SystemExit(f"DB não encontrado: {db_path}")

    if not xlsx_path.exists():
        raise SystemExit(f"Excel não encontrado: {xlsx_path}")

    con = sqlite3.connect(str(db_path))
    try:
        ensure_schema(con)

        wb = load_workbook(str(xlsx_path), data_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        if not rows or len(rows) < 2:
            raise SystemExit("Excel vazio ou sem linhas suficientes.")

        headers = [str(h or "").strip() for h in rows[0]]

        col_setor = pick_col(headers, ["setor", "secao", "seção", "area", "obra", "local"])
        col_nome = pick_col(headers, ["nome", "colaborador", "responsavel", "encarregado"])
        col_funcao = pick_col(headers, ["funcao", "função", "cargo"])
        col_tel = pick_col(headers, ["telefone", "celular", "whatsapp", "fone"])

        if col_nome is None:
            raise SystemExit(
                f"Não achei coluna de NOME. Cabeçalhos encontrados: {headers}"
            )

        inserted = 0
        skipped = 0

        for r in rows[1:]:
            vals = list(r)

            nome = (str(vals[col_nome]).strip() if col_nome is not None and vals[col_nome] is not None else "")
            if not nome:
                continue

            setor = (
                str(vals[col_setor]).strip()
                if col_setor is not None and vals[col_setor] is not None
                else "SEM SETOR"
            )

            funcao = (
                str(vals[col_funcao]).strip()
                if col_funcao is not None and vals[col_funcao] is not None
                else None
            )

            telefone = (
                str(vals[col_tel]).strip()
                if col_tel is not None and vals[col_tel] is not None
                else None
            )

            setor_id = get_or_create_setor(con, setor)

            if exists_encarregado(con, setor_id, nome):
                skipped += 1
                continue

            con.execute(
                "INSERT INTO encarregados (setor_id, funcao, nome, telefone) VALUES (?, ?, ?, ?)",
                (setor_id, funcao, nome, telefone),
            )
            inserted += 1

        con.commit()

        cur = con.cursor()
        total = cur.execute("SELECT COUNT(1) FROM encarregados").fetchone()[0]
        print(f"OK ✅ Import finalizado. Inseridos={inserted} | Duplicados ignorados={skipped} | Total encarregados={total}")
    finally:
        con.close()


if __name__ == "__main__":
    main()
