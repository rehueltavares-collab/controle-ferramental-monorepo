import os
from typing import Dict, List, Tuple

from dotenv import load_dotenv
from openpyxl import load_workbook
import pymysql

# =========================
# CONFIG
# =========================
MONOREPO_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
XLSX_PATH = os.path.join(MONOREPO_ROOT, "planejamento_kits_v4.xlsx")

SHEET_NAME = None  # None = primeira aba

COL_KIT_NOME = "kit_nome"
COL_SETOR = "setor"
COL_ITEM_PATRIMONIO = "patrimonio"
COL_ITEM_DESCRICAO = "descricao"

DRY_RUN = False  # Execucao real

# =========================
# DB
# =========================
def db_conn():
    # Prefer .env na raiz, mas suporta backend/.env se existir
    root_env = os.path.join(MONOREPO_ROOT, ".env")
    backend_env = os.path.join(MONOREPO_ROOT, "backend", ".env")
    if os.path.exists(root_env):
        load_dotenv(root_env)
    if os.path.exists(backend_env):
        load_dotenv(backend_env, override=True)
    host = os.getenv("DB_HOST", "localhost")
    user = os.getenv("DB_USER", "root")
    pwd = os.getenv("DB_PASSWORD", "")
    port = int(os.getenv("DB_PORT", "3306"))
    name = os.getenv("DB_NAME", "ferramental")
    return pymysql.connect(
        host=host,
        user=user,
        password=pwd,
        database=name,
        port=port,
        charset="utf8mb4",
        cursorclass=pymysql.cursors.DictCursor,
        autocommit=False,
    )


# =========================
# XLSX
# =========================
def load_rows_xlsx(path: str) -> List[dict]:
    if not os.path.exists(path):
        raise FileNotFoundError(f"Arquivo nao encontrado: {path}")

    wb = load_workbook(path)
    ws = wb[SHEET_NAME] if SHEET_NAME else wb.worksheets[0]

    header = []
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i == 1:
            header = [str(c).strip() if c is not None else "" for c in row]
            continue
        if all(c is None or str(c).strip() == "" for c in row):
            continue
        obj = {}
        for idx, key in enumerate(header):
            if not key:
                continue
            obj[key.strip()] = row[idx]
        rows.append(obj)
    return rows


def norm(s) -> str:
    return ("" if s is None else str(s)).strip()


# =========================
# CLEANUP
# =========================
def cleanup_operational(cur):
    statements = [
        "DELETE FROM solicitacao_itens",
        "DELETE FROM solicitacoes_retirada",
        "DELETE FROM solicitacoes_operacao",
        "DELETE FROM posses",
        "DELETE FROM checklists_semanais",
        "DELETE FROM item_movimentos",
        "DELETE FROM movimentos",
        "DELETE FROM kit_itens",
        "DELETE FROM kits",
        "DELETE FROM termos_responsabilidade",
    ]
    for sql in statements:
        print(f"[CLEAN] {sql}")
        if not DRY_RUN:
            cur.execute(sql)


# =========================
# LOOKUPS
# =========================
def find_item_id_by_patrimonio(cur, patrimonio: str) -> int:
    cur.execute("SELECT id FROM itens WHERE patrimonio = %s LIMIT 1", (patrimonio,))
    r = cur.fetchone()
    return int(r["id"]) if r else 0


def create_kit(cur, nome: str, setor_id: int = None) -> int:
    cur.execute("SHOW COLUMNS FROM kits")
    cols = {row["Field"] for row in cur.fetchall()}
    fields = ["nome"]
    values = [nome]
    if "setor_id" in cols and setor_id:
        fields.append("setor_id")
        values.append(setor_id)
    if "ativo" in cols:
        fields.append("ativo")
        values.append(1)
    if "disponivel" in cols:
        fields.append("disponivel")
        values.append(1)
    placeholders = ",".join(["%s"] * len(fields))
    cur.execute(
        f"INSERT INTO kits ({', '.join(fields)}) VALUES ({placeholders})",
        tuple(values),
    )
    return int(cur.lastrowid)


def upsert_kit_item(cur, kit_id: int, item_id: int):
    cur.execute(
        "SELECT 1 FROM kit_itens WHERE kit_id=%s AND item_id=%s LIMIT 1",
        (kit_id, item_id),
    )
    if cur.fetchone():
        return
    cur.execute("SHOW COLUMNS FROM kit_itens")
    cols = {row["Field"] for row in cur.fetchall()}
    if "quantidade" in cols:
        cur.execute(
            "INSERT INTO kit_itens (kit_id, item_id, quantidade) VALUES (%s, %s, %s)",
            (kit_id, item_id, 1),
        )
    else:
        cur.execute("INSERT INTO kit_itens (kit_id, item_id) VALUES (%s, %s)", (kit_id, item_id))


def mark_items_allocated(cur, item_ids: List[int]):
    if not item_ids:
        return
    chunk_size = 500
    for i in range(0, len(item_ids), chunk_size):
        chunk = item_ids[i : i + chunk_size]
        fmt = ",".join(["%s"] * len(chunk))
        cur.execute(f"UPDATE itens SET disponivel=0 WHERE id IN ({fmt})", tuple(chunk))


def mark_items_not_allocated(cur):
    cur.execute(
        """
        UPDATE itens i
        LEFT JOIN kit_itens ki ON ki.item_id = i.id
        SET i.disponivel = 1
        WHERE ki.item_id IS NULL
        """
    )


# =========================
# MAIN
# =========================
def resolve_columns(rows: List[dict]) -> Tuple[str, str, str | None]:
    header = set(rows[0].keys()) if rows else set()
    kit_candidates = ["kit_nome", "destino", "kit", "nome_kit"]
    pat_candidates = ["patrimonio", "identificacao", "patrimônio", "patrimonio_item"]
    setor_candidates = ["setor", "setor_kit", "setor kit", "setor_nome", "setor_kit_nome"]
    kit_col = next((c for c in kit_candidates if c in header), None)
    pat_col = next((c for c in pat_candidates if c in header), None)
    setor_col = next((c for c in setor_candidates if c in header), None)
    if not kit_col or not pat_col:
        raise RuntimeError(
            f"Colunas obrigatorias nao encontradas. Header: {sorted(header)}"
        )
    return kit_col, pat_col, setor_col


def find_or_create_setor(cur, setor_nome: str) -> int:
    nome = (setor_nome or "").strip() or "GERAL"
    cur.execute("SELECT id FROM setores WHERE nome = %s LIMIT 1", (nome,))
    row = cur.fetchone()
    if row:
        return int(row["id"])
    cur.execute("INSERT INTO setores (nome) VALUES (%s)", (nome,))
    return int(cur.lastrowid)


def is_avulso_kit_name(nome: str) -> bool:
    return "AVULSO" in (nome or "").upper()


def main():
    rows = load_rows_xlsx(XLSX_PATH)
    if not rows:
        raise RuntimeError("XLSX vazio.")

    kit_col, pat_col, setor_col = resolve_columns(rows)

    kits_map: Dict[str, List[dict]] = {}
    for r in rows:
        kit_nome = norm(r.get(kit_col))
        pat = norm(r.get(pat_col))
        if not kit_nome or not pat:
            continue
        if is_avulso_kit_name(kit_nome):
            continue
        kits_map.setdefault(kit_nome, []).append(r)

    print(f"[INFO] Kits no XLSX: {len(kits_map)}")
    print(f"[INFO] DRY_RUN = {DRY_RUN}")

    conn = db_conn()
    try:
        with conn.cursor() as cur:
            cleanup_operational(cur)

            created_kits: Dict[str, int] = {}
            allocated_item_ids: List[int] = []
            missing_items: List[Tuple[str, str]] = []

            for kit_nome, kit_rows in kits_map.items():
                print(f"[KIT] {kit_nome} (itens: {len(kit_rows)})")
                setor_id = None
                if setor_col:
                    setor_nome = norm(kit_rows[0].get(setor_col))
                    if not DRY_RUN:
                        setor_id = find_or_create_setor(cur, setor_nome)
                kit_id = -1 if DRY_RUN else create_kit(cur, kit_nome, setor_id=setor_id)
                created_kits[kit_nome] = kit_id

                for rr in kit_rows:
                    pat = norm(rr.get(pat_col))
                    if not pat:
                        continue
                    item_id = 1 if DRY_RUN else find_item_id_by_patrimonio(cur, pat)
                    if not item_id and not DRY_RUN:
                        missing_items.append((kit_nome, pat))
                        continue
                    if not DRY_RUN:
                        upsert_kit_item(cur, kit_id, item_id)
                        allocated_item_ids.append(item_id)

            print("[STEP] Reclassificando disponibilidade")
            if not DRY_RUN:
                mark_items_allocated(cur, allocated_item_ids)
                mark_items_not_allocated(cur)

            if missing_items:
                print("\n[ALERTA] Patrimonios nao encontrados em 'itens':")
                for kit_nome, pat in missing_items[:50]:
                    print(f"  * {kit_nome} -> {pat}")
                print(f"[ALERTA] Total faltantes: {len(missing_items)}")

            if DRY_RUN:
                print("\n[DRY_RUN] Nada foi gravado no banco.")
                conn.rollback()
            else:
                conn.commit()
                print("\n[OK] Migracao concluida.")
    finally:
        conn.close()


if __name__ == "__main__":
    main()
